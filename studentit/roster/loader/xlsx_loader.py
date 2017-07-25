import os
import re
import errno
from pathlib import Path

import dateutil.parser
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, coordinate_from_string, column_index_from_string
from dateutil import parser

from studentit.roster.shift import Shift
from ..roster import Roster
from .roster_loader import RosterLoader


class ShiftCollector(object):
    def __init__(self, name, library_colour_index, cell_time_index):
        self.name = name
        self._library_colour_index = library_colour_index
        self._cell_time_index = cell_time_index

        self.shifts = []

    def add_partial(self, colour, cell_row, text):
        location = self._library_colour_index.get(colour)
        start_time, end_time = self._cell_time_index.get(cell_row)

        if location is None:
            raise Exception('Unknown colour location')

        if len(self.shifts) == 0 or self.shifts[-1].location != location:
            self.shifts.append(Shift(name=self.name, location=location, start_time=start_time, end_time=end_time))

        self.shifts[-1].end_time = end_time


class XlsxLoader(RosterLoader):
    def __init__(self, config):
        super().__init__()

        self.config = config

    def load(self, path):
        self.logger.debug('Loading roster from file at {}'.format(path))
        if not Path(path).is_file():
            raise FileNotFoundError(errno.ENOENT, os.strerror(errno.ENOENT), path)

        ws = self._sheet(path)
        start_date, end_date = self._date_range(ws)
        roster = Roster(start_date=start_date, end_date=end_date)

        people = self._people(ws)
        cell_time_index = self._cell_time_index(ws)
        library_colour_index = self._library_colour_index(ws)

        people_col_num, people_start_row = self._split_coords(self.config['people_start_cell'])
        for i in range(0, len(people)):
            people_col = get_column_letter(people_col_num)
            name = ws['{}{}'.format(people_col, people_start_row)].value

            self.logger.debug('Parsing shifts for {}'.format(name))

            blank_in_a_row = 0
            shift_collector = ShiftCollector(name, library_colour_index, cell_time_index)
            for shift_row in range(people_start_row + 1, self.config['max_row']):
                cell_ref = f'{people_col}{shift_row}'
                cell = ws[cell_ref]

                if blank_in_a_row > self.config['blank_limit']:
                    self.logger.debug(f'Reached blank limit ({self.config["blank_limit"]}) for {name} at {cell_ref}. '
                                      f'Assuming there are no more shifts to parse for this staff member.')
                    break

                # Ignore blank cells - only coloured cells are shifts
                if cell.fill.start_color.index == self.config['blank_colour']:
                    blank_in_a_row += 1
                    continue
                else:
                    blank_in_a_row = 0

                shift_collector.add_partial(cell.fill.start_color.index, shift_row, cell.value)

            self.logger.info(f'Collected {len(shift_collector.shifts)} shifts for {name}')

            for shift in shift_collector.shifts:
                roster.add_shift(shift)

            people_col_num += 1

        self.logger.info(f'Collected {len(roster.shifts)} shifts in total')

        return roster

    def _split_coords(self, cell_ref):
        col, row = coordinate_from_string(cell_ref)
        return column_index_from_string(col), row

    def _people(self, ws):
        self.logger.info('Fetching names of workers from worksheet')
        people = []
        for row in ws['C6:ZZ6']:
            for cell in row:
                if cell.value is None:
                    break
                people.append(cell.value)
        return people

    def _library_colour_index(self, ws):
        self.logger.info('Fetching colours from worksheet')
        index = {}
        for row in ws['C2:D4']:
            name = colour = None
            for cell in row:
                if cell.value is not None:
                    name = cell.value
                if cell.fill.start_color.index is not None and cell.fill.start_color.index != self.config['blank_colour']:
                    colour = cell.fill.start_color.index
            if name is not None and colour is not None:
                index[colour] = name
                self.logger.info(f'Found {name} to be {colour}')
            elif name is not None or colour is not None:
                self.logger.error(f'Found bad location information in {row} with name={name}, colour={colour}')
        return index

    def _cell_time_index(self, ws):
        self.logger.info('Computing which cell references correspond to which dates and times')
        index = {}

        cur_date = None
        for row in ws['A7:B200']:
            for cell in row:
                if cell.value is None:
                    continue

                if '-' not in cell.value:
                    try:
                        date_text = self._clean_date_value(cell.value)
                        cur_date = parser.parse(date_text)
                    except ValueError:
                        self.logger.warning(f'Could not parse potential date cell {date_text}')
                else:
                    time_text = cell.value
                    start_time, end_time = time_text.split(' - ', 2)

                    start_time =  parser.parse(start_time)
                    end_time = parser.parse(end_time)

                    start_time.replace(year=cur_date.year, month=cur_date.month, day=cur_date.day)
                    end_time.replace(year=cur_date.year, month=cur_date.month, day=cur_date.day)

                    index[cell.row] = start_time, end_time

        return index

    def _clean_date_value(self, value):
        # Fix common misspellings for Tuesday which we cannot parse
        value = value.replace('Tues', 'Tue')

        # Fix common misspellings for Thursday which we cannot parse
        value = value.replace('Thurs', 'Thu')
        value = value.replace('Thur', 'Thu')

        return value

    def _sheet(self, path):
        self.logger.debug('Loading workbook')
        wb = load_workbook(path)

        # Always just use the first worksheet
        sheet = wb.worksheets[0]
        self.logger.debug('Using sheet {}'.format(sheet.title))

        return sheet

    def _date_range(self, sheet):
        date_cell_ref = self.config['date_cell']
        cell_format = self.config['date_cell_format']

        self.logger.debug(f'Getting date from {date_cell_ref} using format {cell_format}')

        date_cell_text = sheet[date_cell_ref].value

        if date_cell_text is None:
            raise Exception('No text in date cell ({}). Check if cell reference is correct in the config file.'.format(
                date_cell_ref))

        matches = re.findall(cell_format, date_cell_text)
        if len(matches) != 1:
            raise Exception('Invalid text format in date cell ({}). Text: {}'.format(date_cell_ref, date_cell_text))

        dates = matches[0]
        if len(dates) != 2:
            raise Exception('Invalid text format in date cell ({}). Text: {}'.format(date_cell_ref, date_cell_text))

        start_date = dateutil.parser.parse(dates[0])
        end_date = dateutil.parser.parse(dates[1])
        end_date = end_date.replace(hour=23, minute=59, second=59)

        self.logger.info('Roster spanning {} - {}'.format(start_date, end_date))

        return start_date, end_date
